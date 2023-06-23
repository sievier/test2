from DataLoad import ClickStreamData
from layer import TPEncoder
from config import conf
import numpy as np
import torch
from torch.autograd import Variable
import torch.utils.data as Data
from torch.nn.utils.rnn import pad_sequence
from DataLoad import TPData
from model import User_Candidate_Encoder
import time
import torch.optim as optim
import math
import os
import xlwt
import openpyxl
import shutil
from tensorboardX import SummaryWriter
from random import sample
from constrain import negtive_log
from collections import Counter
import random
from torchsummary import summary
cfg=conf()
Dir=os.path.abspath('')
model = User_Candidate_Encoder(cfg)
device ='cpu'    #torch.device("cuda:0" if torch.cuda.is_available() else "cpu")
#####**TP信息**###
TPInfor = TPData.TP
ItemIDList = TPInfor[0]  # ItemID索引
# print ('ItemIDList:',ItemIDList)
TitleWordIDlist = TPInfor[1]
# print('TitleWordIDlist:',TitleWordIDlist.shape,TitleWordIDlist[0])
DestinationIDList = TPInfor[2]
# print('DestinationIDList:',DestinationIDList)
Travel_RegionIDList = TPInfor[3]
# print(np.max(Travel_RegionIDList))
# print('Travel_RegionIDList:',Travel_RegionIDList)
Travel_TypeIDList = TPInfor[4]

def test(epoch,month):
    starttime = time.time()
    test_model = torch.load(Dir+'/model/'+'model-'+epoch+month+'.pkl')
    test_model.eval()
    #BATCH_SIZE = cfg.batch_size  # 批训练的数据个数
    BATCH_SIZE=1 #依次计算测试集中的样本
    result=ClickStreamData.ClickStream(month)
    Test_Clickdata=np.array(result[1]) #测试集
    #print (Test_Clickdata)
    ID_traing = np.arange(Test_Clickdata.shape[0])
    UserID_train = Test_Clickdata[:, 0]    #UserID
    ItemID_long_train = Test_Clickdata[:, 1]  #long-termd点击流
    sum_first=[]
    for Item in ItemID_long_train:
        #print('Item:',Item)
        if str(Item)=='[]':
            #print('yes')
            sum_first.append(1)
    print('num of Fitst_visit Users:',sum(sum_first))
    ItemID_short_train = Test_Clickdata[:, 2]  #short-term点击流
    Target_traing=Test_Clickdata[:,3]  #下单的ItemID
    #print (Targe_traing)
    x=torch.from_numpy(ID_traing)
    y=torch.Tensor([int(x) for x in Target_traing]) # 转换为list
    tensor_Test = Data.TensorDataset(x,y)   # 将数据封装成Dataset
    # 可使用索引调用数据
    # print ('tensor_data[0]: ', tensor_dataset[0])
    tensor_dataloader = Data.DataLoader(tensor_Test,  # 封装的对象
                                        batch_size=BATCH_SIZE,  # 输出的batchsize
                                        shuffle=True,  # 随机输出
                                        num_workers=0)  # 只有1个进程
    candidate_title = []
    candidate_destination = []
    candidate_Travel_Type = []
    candidate_Travel_RegionID = []
    CandicateItemIDList=result[4]  ##测试集合中的候选项集
    K=2#随机抽样的推荐候选项集
    #book_weight = xlwt.Workbook(encoding="utf-8", style_compression=0)  #统计view 和长短期兴趣的权重
    book_weight = openpyxl.Workbook()  #统计view 和长短期兴趣的权重
    #sheet_view = book_weight.add_sheet('ViewLevelAttention', cell_overwrite_ok=True)
    #sheet_longshort = book_weight.add_sheet('LongAndShortAttention', cell_overwrite_ok=True)
    sheet_view=book_weight.create_sheet('ViewLevelAttention',index=0)
    sheet_longshort =book_weight.create_sheet('LongAndShortAttention',index=1)
    View_line=1
    sheet_view.cell(View_line, 1).value = 'title'  # 写文件
    sheet_view.cell(View_line, 2).value = 'destination'  # 写文件
    sheet_view.cell(View_line, 3).value = 'travel_type'  # 写文件
    sheet_view.cell(View_line, 4).value = 'travel_region'  # 写文件
    sheet_longshort.cell(1,1).value='Long-term Behaviors'
    sheet_longshort.cell(1,2).value='Short-term Behaviors'
    for i in range(0, len(CandicateItemIDList)):  #计算一遍候选项集(所有的产品)
        ItemID = CandicateItemIDList[i]
        # print('推荐候选ItemID：',ItemID)
        #print ('TitleWordIDlist[i]:',TitleWordIDlist[i])
        candidate_title.append(TitleWordIDlist[i])
        candidate_destination.append([int(DestinationIDList[i])])
        candidate_Travel_Type.append([int(Travel_TypeIDList[i])])
        candidate_Travel_RegionID.append([int(Travel_RegionIDList[i])])
    for step, (batch_data, batch_target) in enumerate(tensor_dataloader):#一个batch（64个会话）
        #batch_data=batch_data.to(device)
        print('******One Batch****')
        #print(data, target)
        #TPEnocderone = TPEncoder(cfg)
        print('|Step:', step, '|batch_data_user:',[UserID_train[int(ID)] for ID in batch_data], '|batch_target', batch_target)
        batch_userID=[]
        batch_long_term_TitleID=[]  #batch_long_term_embedding的输入
        batch_long_term_DestinationID=[]
        batch_long_term_TravelRegionID=[]
        batch_long_term_TravelTypeID=[]
        batch_long_termID=[]
        batch_short_term_TitleID=[]  #batch_short_term_embedding的输入
        batch_short_term_DestinationID = []
        batch_short_term_TravelRegionID = []
        batch_short_term_TravelTypeID = []
        batch_short_termID=[]
        for ID in batch_data:  #该batch中的一个用户（会话）
            #print('ID:',int(ID))
            #print('*****UserID******:',UserID_train[int(ID)])
            batch_userID.append(UserID_train[int(ID)])  #batch——UserID的输入
            #print('Long-term:', ItemID_long_train[int(ID)]) #Long-term点击流
            Long_term_TitleID=[]
            Long_term_DestinationID=[]
            Long_term_TravelRegionID=[]
            Long_term_TravelTypeID=[]
            Long_term_ItemID=[]
            if ItemID_long_train[int(ID)]!=[]:
                for ItemID in ItemID_long_train[int(ID)]:  #一个产品
                    Long_term_ItemID.append(ItemID)
                    #print('ItemID:',ItemID)
                    Index_ID=np.where(ItemIDList == ItemID)
                    WordIDList=TitleWordIDlist[Index_ID]
                    #print(WordIDList)
                    #print ('Title_WordID:',TitleWordIDlist[Index_ID])
                    Long_term_TitleID.append(torch.Tensor(WordIDList[0]))
                    #print(DestinationIDList[Index_ID],Travel_RegionIDList[Index_ID],Travel_TypeIDList[Index_ID])
                    Long_term_DestinationID.append(torch.Tensor(DestinationIDList[Index_ID]))
                    Long_term_TravelRegionID.append(torch.Tensor(Travel_RegionIDList[Index_ID]))
                    Long_term_TravelTypeID.append(torch.Tensor(Travel_TypeIDList[Index_ID]))
                #print('Long_term_TitleID:', Long_term_TitleID)

            #print ('Long_term_TitleID:',Long_term_TitleID)
            else:  #不存在长期行为，全部补0
                Long_term_TitleID.append(torch.Tensor(np.zeros(23)))
                Long_term_DestinationID.append(torch.Tensor(np.zeros(1)))
                Long_term_TravelRegionID.append(torch.Tensor(np.zeros(1)))
                Long_term_TravelTypeID.append(torch.Tensor(np.zeros(1)))
            Long_term_TitleID = pad_sequence(Long_term_TitleID, batch_first=True)#补齐
            #print('Long_term_TitleID.size：', Long_term_TitleID.size())

            Short_term_TitleID = []
            Short_term_DestinationID=[]
            Short_term_TravelRegionID=[]
            Short_term_TravelTypeID=[]
            Short_term_ItemID=[]
            for ItemID in ItemID_short_train[int(ID)]:
                Short_term_ItemID.append(ItemID)
                Index_ID = np.where(ItemIDList == ItemID)
                #print ('ItemID',ItemID,'Index_ID',Index_ID)
                WordIDList = TitleWordIDlist[Index_ID]
                #print(WordIDList)
                Short_term_TitleID.append(torch.Tensor(WordIDList[0]))
                Short_term_DestinationID.append(torch.Tensor(DestinationIDList[Index_ID]))
                Short_term_TravelRegionID.append(torch.Tensor(Travel_RegionIDList[Index_ID]))
                Short_term_TravelTypeID.append(torch.Tensor(Travel_TypeIDList[Index_ID]))
            Short_term_TitleID = pad_sequence(Short_term_TitleID, batch_first=True)
            #print('Short_term_TitleID:', Short_term_TitleID)
            #long_term的4个输入

            batch_long_term_TitleID.append(Long_term_TitleID)
            #print('Long_term_DestinationID:', Long_term_DestinationID)
            batch_long_term_DestinationID.append(torch.Tensor(Long_term_DestinationID))
            #print('Long_term_TravelRegionID:',Long_term_TravelRegionID)
            batch_long_term_TravelRegionID.append(torch.Tensor(Long_term_TravelRegionID))
            #print('Long_term_TravelTypeID:',Long_term_TravelTypeID)
            batch_long_term_TravelTypeID.append(torch.Tensor(Long_term_TravelTypeID))
            batch_long_termID.append(Long_term_ItemID)
            #short_term的4个输入
            batch_short_term_TitleID.append(Short_term_TitleID)
            batch_short_term_DestinationID.append(torch.Tensor(Short_term_DestinationID))
            batch_short_term_TravelRegionID.append(torch.Tensor(Short_term_TravelRegionID))
            batch_short_term_TravelTypeID.append(torch.Tensor(Short_term_TravelTypeID))
            batch_short_termID.append(Short_term_ItemID)

        #print('得出一个batch的输出：')
        batch_userID=np.array(batch_userID)
        x1=torch.from_numpy(batch_userID)

        #print('batch_userID:',x1.size())
        #print('batch_userID:',batch_userID.shape,batch_userID)  #UserID
        #print('batch_long_term_TitleID:',batch_long_term_TitleID)
        x2=pad_sequence(batch_long_term_TitleID, batch_first=True)    #x2,y2:Title;
        x2= torch._cast_Long(x2)
        #print('x2:',x2.size())
        #print('batch_long_term_Title_embedding：',x2.size())

        x3=pad_sequence(batch_long_term_DestinationID, batch_first=True)  #x3,y3:Destination;
        x3 = torch._cast_Long(x3)
        #print('batch_long_term_DestinationID:',x3.size())
        x4=pad_sequence(batch_long_term_TravelTypeID, batch_first=True)  #x4,y4:TravelType;
        x4 = torch._cast_Long(x4)
        #print('batch_long_term_TravelRegionID:', x4.size())
        x5=pad_sequence(batch_long_term_TravelRegionID,batch_first=True) #x5,y5:TravelRegionID;
        x5 = torch._cast_Long(x5)
        #print('batch_long_term_TravelTypeID:', x5.size())

        y1=x1
        #print('batch_userID:', y1.size)
        y2=pad_sequence(batch_short_term_TitleID, batch_first=True)
        y2 = torch._cast_Long(y2)
        #print('batch_short_term_Title_embedding：',y2.size())
        y3=pad_sequence(batch_short_term_DestinationID, batch_first=True)
        y3 = torch._cast_Long(y3)
        #print('batch_short_term_DestinationID:',y3.size())
        y4=pad_sequence(batch_short_term_TravelTypeID, batch_first=True)
        y4 = torch._cast_Long(y4)
        #print('batch_short_term_TravelRegionID:', y4.size())
        y5 = pad_sequence(batch_short_term_TravelRegionID, batch_first=True)
        y5 = torch._cast_Long(y5)
        #print('batch_short_term_TravelTypeID:', y5.size())
        #print('batch_target:',batch_target.shape,batch_target)  #Odered ItemOD
        print('END Data Loading：')
        ########******一个batch,Start to Training******####
        print("Start to UserEncoder:")
        #print(model)
        x1 = torch.unsqueeze(x1, dim=1)
        x1 = torch.unsqueeze(x1, dim=1)
        x1 = x1.expand(-1, x2.size(1), -1)
        x3 = torch.unsqueeze(x3, dim=-1)
        x4 = torch.unsqueeze(x4, dim=-1)
        x5 = torch.unsqueeze(x5, dim=-1)

        y1 = torch.unsqueeze(y1, dim=1)
        y1 = torch.unsqueeze(y1, dim=1)
        y1 = y1.expand(-1, y2.size(1), -1)
        y3 = torch.unsqueeze(y3, dim=-1)
        y4 = torch.unsqueeze(y4, dim=-1)
        y5 = torch.unsqueeze(y5, dim=-1)
        #print('Start negativesampling for candidateItem!')  # 消极样本采样
        p1=batch_target #user下单的ItemID
        #print('batch_positive_label',p1)
        s_title=[] #positive_sample_title
        s_destination=[]#positive_destination
        s_travel_type=[]#positive_Travel_Type
        s_travel_region=[]#positive_Travel_RegionID
        CandicateItemIDList_sample=CandicateItemIDList
        random.shuffle(CandicateItemIDList)
        CandicateItemIDList_sample = CandicateItemIDList_sample[:K]  #采样K个候选项集
        CandicateItemIDList_sample.extend(Long_term_ItemID)  #加入Long-term点击流
        CandicateItemIDList_sample.extend(Short_term_ItemID)#加入short-term点击流

        for itemID in CandicateItemIDList_sample:  #一个batch下的下单ItemID
            #print('itemID:',int(itemID))
            #推荐候选项集的特征
            index_p=CandicateItemIDList.index(int(itemID))
            s_title.append(candidate_title[index_p])
            s_destination.append(candidate_destination[index_p])
            s_travel_type.append(candidate_Travel_Type[index_p])
            s_travel_region.append(candidate_Travel_RegionID[index_p])
        s_title=torch.tensor([s_title])
        s_destination= torch.tensor([s_destination])
        s_travel_type= torch.tensor([s_travel_type])
        s_travel_region= torch.tensor([s_travel_region])
        # s_title=s_title[:,0:100,:]
        # s_destination=s_destination[:,0:100,:]
        # s_travel_type=s_travel_type[:,0:100,:]
        # s_travel_region=s_travel_region[:,0:100,:]
        #print('p2,p3,p4,p5:',s_title.size(),s_destination.size(),s_travel_type.size(),s_travel_region.size())
        encoder=test_model(x1,x2,x3,x4,x5,y1,y2,y3,y4,y5,s_title,s_destination,s_travel_type,s_travel_region)
        batch_userbehaviorencoder=torch.tensor(encoder[0]) #batch_user Encoder
        batch_sample_Encoder=encoder[1]  #
        batch_weight_view=encoder[2]  #batch_sample_view-level weight
        batch_Long_Weight=encoder[3] #Fu_ weight of Long-term behavior
        batch_Short_Weight=1-batch_Long_Weight #Short of Long-term behavior
        #print('batch_sample_view-level weight:',batch_weight_view)  # x2,x3,x4,x5分别对应s_title,s_destination,s_travel_type,s_travel_region
        for pair in batch_weight_view:
            #print('pair:', pair[0])
            title=pair[0,0]
            destination=pair[0,1]
            travel_type=pair[0,2]
            travel_region=pair[0,3]
            View_line=View_line+1
            try:
                # sheet_view.write(View_line,0,str(title[0]))
                # sheet_view.write(View_line,1,str(destination[0]))
                # sheet_view.write(View_line,2,str(travel_type[0]))
                # sheet_view.write(View_line,3,str(travel_region[0]))
                sheet_view.cell(View_line, 1).value = float(title[0])  # 写文件
                sheet_view.cell(View_line, 2).value = float(destination[0])  # 写文件
                sheet_view.cell(View_line, 3).value = float(travel_type[0])  # 写文件
                sheet_view.cell(View_line, 4).value = float(travel_region[0])  # 写文件
            except Exception as e:
                print(title[0], destination[0], travel_type[0], travel_region[0])
                print(e)

        #print('batch_Long_Weight:',batch_Long_Weight[0,0],batch_Long_Weight.item())
        #print('batch_Short_Weight:',batch_Short_Weight[0,0],batch_Short_Weight.item())
        # sheet_longshort.write(step+1,0,batch_Long_Weight.item())
        # sheet_longshort.write(step+1,1, batch_Short_Weight.item())
        sheet_longshort.cell(step+2, 1).value = float(batch_Long_Weight.item())  # 写文件
        sheet_longshort.cell(step+2, 2).value = float(batch_Short_Weight.item())  # 写文件
        book_weight.save(Dir+'\\experiment\\Visualization Attention'+month+'.xlsx')
        batch_userbehaviorencoder = torch.unsqueeze(batch_userbehaviorencoder, dim=1)
        print('batch_userbehaviorencoder_shape:', batch_userbehaviorencoder.size())
        print('batch_sample_Encoder_shape:', batch_sample_Encoder.size())  #所有的推荐候选项集
        batch_Score = torch.sigmoid(torch.sum(torch.mul(batch_userbehaviorencoder,batch_sample_Encoder), dim=-1))  # torch.Size([2, 1])
        #print(batch_Score.size())
    endtime = time.time()
    running_time = endtime - starttime
    print('Running Time:', running_time / 60.0, '分')
test('0','07')
#test('0','08')
