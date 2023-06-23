#!/usr/bin/env python
# -*- coding: utf-8 -*-
import os
import xlrd
import xlwt
import numpy as np
import pandas as pd
import datetime
import openpyxl
import time
from sklearn.model_selection import train_test_split
from gensim import corpora
Dir=os.path.abspath('')
starttime = time.time()

def dataload(month):
    filename = 'ClickStream' + month + '.xls'
    data = xlrd.open_workbook('C:\\Users\\USER\\OneDrive\\Deep Learning\\Tuniu-master\\data\\'+ filename)
    table = data.sheets()[0]  # 通过索引顺序获取
    UserIDlist = []
    Training_Userdata = []  # 用1-29号下单数据训练
    Test_Userdata = []  # 用20和31号下单数据测试预测
    OrderItemIDList = []
    book_clickstream = openpyxl.Workbook()

    sheet =book_clickstream.create_sheet(index=0)
    sheet.cell(1, 1).value='user_id' # UserID
    sheet.cell(1, 2).value='item_id'  # UserID
    sheet.cell(1, 3).value='rating'  # UserID
    sheet.cell(1, 4).value='label'  # UserI
    j=2
    ItemList=[]
    for i in range(0, table.nrows):
        # print (table.row(i))
        UserID_i = int(table.cell_value(i, 0))
        UserIDlist.append(UserID_i)
        # print ("UserID_i:",UserID_i)
        ClickItemID_i = table.cell_value(i, 1).strip('\n')
        if '|' in ClickItemID_i:
            ClickItemID_i_list = ClickItemID_i.split('|')
        else:
            ClickItemID_i_list = [int(ClickItemID_i)]
        ClickItemID_i_list = [int(x) for x in ClickItemID_i_list]
        OrderItemID_i = int(table.cell_value(i, 4))
        ClickItemID_i_list.append(OrderItemID_i)
        ItemList.extend(ClickItemID_i_list)
    ItemIDList=list(set(ItemList))
    #print ('ItemList:',ItemList)
    for i in range(0, table.nrows):
        # print (table.row(i))
        UserID_i = int(table.cell_value(i, 0))
        UserIDlist.append(UserID_i)
        # print ("UserID_i:",UserID_i)
        ClickItemID_i = table.cell_value(i, 1).strip('\n')
        if '|' in ClickItemID_i:
            ClickItemID_i_list = ClickItemID_i.split('|')
        else:
            ClickItemID_i_list = [int(ClickItemID_i)]
        ClickItemID_i_list = [int(x) for x in ClickItemID_i_list]
        disClickItemID_i_list=list(set(ClickItemID_i_list))
        OrderItemID_i = int(table.cell_value(i, 4))
        for Item in disClickItemID_i_list:
            if Item!=OrderItemID_i:   #剔除买的Item
                sheet.cell(j, 1).value=UserID_i # UserID
                sheet.cell(j, 2).value=Item
                sheet.cell(j, 3).value= ClickItemID_i_list.count(Item)  # UserID
                sheet.cell(j, 4).value=OrderItemID_i # UserI
                j=j+1
    book_clickstream.save(Dir+'\\User-Item-Clicks-Label'+month+'.xls')

def get_key(dct, value):
    key=[k for (k,v) in dct.items() if v == value]
    return key[0]

month='07'
dataload(month)  #生成文件

print('Data Loading Finish!')
header = ['user_id', 'item_id', 'rating', 'label']
df = pd.read_excel(Dir + '\\User-Item-Clicks-Label' + month + '.xls', sep='\t', names=header)
# print(df)


UserID = df.user_id.tolist()
ItemID = df.item_id.tolist()
ItemID_MAP = dict(enumerate(list(set(ItemID))))
# print (len(ItemID))
# print(len(list(set(ItemID))))
rating = df.rating.tolist()
label = df.label.tolist()
ItemID1 = df.item_id.tolist()
ItemID1.extend(label)

UserID_MAP = dict(enumerate(list(set(UserID))))

ItemID1_MAP = dict(enumerate(list(set(ItemID1))))
# print(ItemID1_MAP
print('END MAP!')


# print(UserID_MAP,ItemID_MAP)

##print (len(ItemID))

ItemID_MAP_list = [get_key(ItemID1_MAP, value) for value in ItemID]
print('ItemID:', np.array(ItemID_MAP_list).shape)

UserID_MAP_list = [get_key(UserID_MAP, value) for value in UserID]  # 转换为MAPID
print('UserID:', np.array(UserID_MAP_list).shape)

Label_MAP_list = [get_key(ItemID1_MAP, value) for value in label]
print('Label:', np.array(Label_MAP_list).shape)

df = pd.DataFrame({"user_id": UserID_MAP_list, 'item_id': ItemID_MAP_list, 'rating': rating, 'label': Label_MAP_list})
df.to_csv(Dir + '\\UserMap-ItemMap-Clicks-Label' + month + '.csv', encoding='gbk')
print(df)
print('数据MAP结束，开始计算:')


header = ['user_id', 'item_id', 'rating', 'label']
df = pd.read_csv(Dir+'\\UserMap-ItemMap-Clicks-Label'+month+'.csv', header=0,names=header)
print (df)

# 计算唯一用户和电影的数量
n_users = df.user_id.unique().shape[0]
n_items = df.item_id.unique().shape[0]
n_items = len(ItemID1_MAP)
print('Number of users = ' + str(n_users) + ' | Number of travel packages = ' + str(n_items))

train_data, test_data = train_test_split(df, test_size=1, random_state=21)

# 协同过滤算法
# 第一步是创建uesr-item矩阵，此处需创建训练和测试两个UI矩阵
train_data_matrix = np.zeros((n_users, n_items))
for line in train_data.itertuples():
    train_data_matrix[line[1] - 1, line[2] - 1] = line[3]

test_data_matrix = np.zeros((n_users, n_items))
for line in test_data.itertuples():
    test_data_matrix[line[1] - 1, line[2] - 1] = line[3]

print(train_data_matrix.shape)
print(test_data_matrix.shape)

# # 计算相似度
# # 使用sklearn的pairwise_distances函数来计算余弦相似性
# from sklearn.metrics.pairwise import pairwise_distances
# # 计算用户相似度
# user_similarity = pairwise_distances(train_data_matrix, metric='cosine')
# # 计算物品相似度
# item_similarity = pairwise_distances(train_data_matrix.T, metric='cosine')

# 计算相似度
# 使用sklearn的cosine_similarity函数来计算余弦相似性
from sklearn.metrics.pairwise import cosine_similarity

# 计算用户相似度
user_similarity = cosine_similarity(train_data_matrix)
# 计算物品相似度
item_similarity = cosine_similarity(train_data_matrix.T)

print(u"用户相似度矩阵维度：", user_similarity.shape, u"  物品相似度矩阵维度：", item_similarity.shape)
#print(u"用户相似度矩阵：", user_similarity)
#print(u"物品相似度矩阵：", item_similarity)

# 预测
def predict(ratings, similarity, type):
    # 基于用户相似度矩阵的
    if type == 'user':
        mean_user_rating = ratings.mean(axis=1)
        # You use np.newaxis so that mean_user_rating has same format as ratings
        ratings_diff = (ratings - mean_user_rating[:, np.newaxis])
        pred = mean_user_rating[:, np.newaxis] + np.dot(similarity, ratings_diff) / np.array(
            [np.abs(similarity).sum(axis=1)]).T
    # 基于物品相似度矩阵的
    elif type == 'item':
        pred = ratings.dot(similarity) / np.array([np.abs(similarity).sum(axis=1)])
    print(u"预测值: ", pred.shape)
    return pred


# 预测结果
top_k=50
book = openpyxl.Workbook()
sheet = book.create_sheet(index=0)
sheet.cell(1, 1).value = 'user_id'  # UserID
sheet.cell(1, 2).value = 'RC_item_id'  # UserID
sheet.cell(1, 3).value = 'label'  # UserI

#user_prediction = predict(train_data_matrix, user_similarity, type='user')
#prediction=user_prediction

item_prediction = predict(train_data_matrix, item_similarity, type='item')
prediction=item_prediction
#print('ICF:',item_prediction)

print('user_prediction:',prediction.shape)
recall=[]
for i in range(0,prediction.shape[0]):
    top_k_idx = prediction[i].argsort()[::-1][0:top_k]
    #print ('user:',i,'************>>>>>>',top_k_idx)
    #print('label:',UsreID_Label[i])
    sheet.cell(i+2,1).value=UserID_MAP[i]
    rc=[ItemID1_MAP[k] for k in top_k_idx]
    sheet.cell(i+2,2).value=str(rc)
    groundtruth=int(label[UserID.index(UserID_MAP[i])])
    if groundtruth in rc:
        recall.append(1)
        print ('user:',UserID_MAP[i],'************>>>>>>',rc)
        print('groundtruth:',groundtruth)
    else:
        recall.append(0)
    sheet.cell(i+2,3).value=groundtruth
print ('Average Recall:',np.mean(recall))
#book_UCF.save(Dir+'\\UCFRC-Label'+month+'.xls')
book.save(Dir+'\\ICFRC-Label'+month+'.xls')

endtime = time.time()
running_time = endtime - starttime
print('Running Time:', running_time / 60.0, '分')




# 评估指标，均方根误差
# 使用sklearn的mean_square_error (MSE)函数，其中，RMSE仅仅是MSE的平方根
# 这里只是想要考虑测试数据集中的预测评分，
# 因此，使用prediction[ground_truth.nonzero()]筛选出预测矩阵中的所有其他元素
'''
from sklearn.metrics import mean_squared_error
from math import sqrt

def rmse(prediction, ground_truth):
    prediction = prediction[ground_truth.nonzero()].flatten()
    ground_truth = ground_truth[ground_truth.nonzero()].flatten()
    return sqrt(mean_squared_error(prediction, ground_truth))


print(train_data_matrix)
print(test_data_matrix)
print('User-based CF RMSE: ' + str(rmse(user_prediction, test_data_matrix)))
item_prediction = np.nan_to_num(item_prediction)
print('Item-based CF RMSE: ' + str(rmse(item_prediction, test_data_matrix)))
'''
# 缺点:没有解决冷启动问题，也就是当新用户或新产品进入系统时。
