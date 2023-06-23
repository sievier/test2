"""
This module illustrates how to retrieve the top-10 items with highest rating
prediction. We first train an SVD algorithm on the MovieLens dataset, and then
predict all the ratings for the pairs (user, item) that are not in the training
set. We then retrieve the top-10 prediction for each user.
"""

from __future__ import (absolute_import, division, print_function,
                        unicode_literals)
from collections import defaultdict
import openpyxl
from surprise import SVD
from surprise import Dataset
import os
from surprise import Reader
import time
starttime = time.time()
Dir=os.path.abspath('')

def get_top_n(predictions, n=50):
    '''Return the top-N recommendation for each user from a set of predictions.

    Args:
        predictions(list of Prediction objects): The list of predictions, as
            returned by the test method of an algorithm.
        n(int): The number of recommendation to output for each user. Default
            is 10.

    Returns:
    A dict where keys are user (raw) ids and values are lists of tuples:
        [(raw item id, rating estimation), ...] of size n.
    '''

    # First map the predictions to each user.
    top_n = defaultdict(list)
    for uid, iid, true_r, est, _ in predictions:
        top_n[uid].append((iid, est))

    # Then sort the predictions for each user and retrieve the k highest ones.
    for uid, user_ratings in top_n.items():
        user_ratings.sort(key=lambda x: x[1], reverse=True)
        top_n[uid] = user_ratings[:n]

    return top_n


# First train an SVD algorithm on the movielens dataset.
#data = Dataset.load_builtin('ml-100k')
def RC(month):
    reader = Reader(line_format='user item rating timestamp', sep='\t')
    # 指定要读入的数据文件，本例中为test.txt
    data = Dataset.load_from_file('C:\\Users\\USER\\OneDrive\\Deep Learning\\Tuniu-master\\Baseline\\UserMap-ItemMap-Clicks-Label'+month+'.txt', reader=reader)

    trainset = data.build_full_trainset()
    algo = SVD() #'lr_all':0.01
    algo.fit(trainset)

    # Than predict ratings for all pairs (u, i) that are NOT in the training set.
    testset = trainset.build_anti_testset()
    predictions = algo.test(testset)

    top_n = get_top_n(predictions, n=50)

    # Print the recommended items for each user
    book_clickstream = openpyxl.Workbook()

    sheet = book_clickstream.create_sheet(index=0)
    sheet.cell(1, 1).value = 'user_id'  # UserID
    sheet.cell(1, 2).value = 'RC_item_id'  # UserID

    j=0
    for uid, user_ratings in top_n.items():
        j=j+1
        print(uid, [iid for (iid, _) in user_ratings])
        sheet.cell(j, 1).value = uid  # UserID
        sheet.cell(j, 2).value = str([iid for (iid, _) in user_ratings])
    book_clickstream.save('C:\\Users\\USER\\OneDrive\\Deep Learning\\Tuniu-master\\Baseline\\SVD RC'+month+'+.xls')

RC('07')

endtime = time.time()
running_time = endtime - starttime
print('Running Time:', running_time / 60.0, '分')